# -*- coding: utf-8 -*-
"""
Export Manager - تصدير البيانات الى Excel و CSV
"""

import csv
import os
from datetime import datetime
from kivy.utils import platform

class ExportManager:
    """Handle data export to various formats"""

    def __init__(self):
        self.export_dir = self._get_export_directory()

    def _get_export_directory(self):
        """Get export directory path"""
        if platform == 'android':
            from android.storage import primary_external_storage_path
            base_dir = primary_external_storage_path()
            export_dir = os.path.join(base_dir, 'CustomerManager', 'Exports')
        else:
            export_dir = os.path.join(os.path.expanduser('~'), 'CustomerManager', 'Exports')

        os.makedirs(export_dir, exist_ok=True)
        return export_dir

    def export_to_csv(self, data, filename=None):
        """Export data to CSV file"""
        if not filename:
            filename = f'customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        filepath = os.path.join(self.export_dir, filename)

        if 'customers' in data and data['customers']:
            customers = data['customers']
            fieldnames = ['id', 'name', 'phone', 'email', 'address', 'service_type',
                         'amount', 'paid_amount', 'remaining_amount', 'start_date',
                         'end_date', 'renewal_date', 'notes', 'status', 'created_at']

            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for customer in customers:
                    writer.writerow({k: customer.get(k, '') for k in fieldnames})

        return filepath

    def export_to_excel(self, data, filename=None):
        """Export data to Excel file using openpyxl"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self.export_to_csv(data, filename.replace('.xlsx', '.csv') if filename else None)

        if not filename:
            filename = f'customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        filepath = os.path.join(self.export_dir, filename)

        wb = openpyxl.Workbook()

        # Customers sheet
        ws_customers = wb.active
        ws_customers.title = 'العملاء'
        ws_customers.sheet_view.rightToLeft = True

        if 'customers' in data and data['customers']:
            customers = data['customers']
            headers = ['الرقم', 'الاسم', 'الهاتف', 'البريد', 'العنوان', 'نوع الخدمة',
                      'المبلغ', 'المدفوع', 'المتبقي', 'تاريخ البدء', 'تاريخ الانتهاء',
                      'تاريخ التجديد', 'ملاحظات', 'الحالة', 'تاريخ الاضافة']

            # Header style
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws_customers.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data
            for row, customer in enumerate(customers, 2):
                values = [
                    customer.get('id', ''),
                    customer.get('name', ''),
                    customer.get('phone', ''),
                    customer.get('email', ''),
                    customer.get('address', ''),
                    customer.get('service_type', ''),
                    customer.get('amount', 0),
                    customer.get('paid_amount', 0),
                    customer.get('remaining_amount', 0),
                    customer.get('start_date', ''),
                    customer.get('end_date', ''),
                    customer.get('renewal_date', ''),
                    customer.get('notes', ''),
                    customer.get('status', ''),
                    customer.get('created_at', '')
                ]

                for col, value in enumerate(values, 1):
                    cell = ws_customers.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for row in range(1, len(customers) + 2):
                    cell = ws_customers[f'{column}{row}']
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_customers.column_dimensions[column].width = adjusted_width

        # Payments sheet
        ws_payments = wb.create_sheet('المدفوعات')
        ws_payments.sheet_view.rightToLeft = True

        if 'payments' in data and data['payments']:
            payments = data['payments']
            payment_headers = ['الرقم', 'رقم العميل', 'المبلغ', 'تاريخ الدفع', 'طريقة الدفع', 'ملاحظات']

            for col, header in enumerate(payment_headers, 1):
                cell = ws_payments.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF', size=12)
                cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row, payment in enumerate(payments, 2):
                values = [
                    payment.get('id', ''),
                    payment.get('customer_id', ''),
                    payment.get('amount', 0),
                    payment.get('payment_date', ''),
                    payment.get('payment_method', ''),
                    payment.get('notes', '')
                ]

                for col, value in enumerate(values, 1):
                    cell = ws_payments.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            for col in range(1, len(payment_headers) + 1):
                column = get_column_letter(col)
                ws_payments.column_dimensions[column].width = 20

        wb.save(filepath)
        return filepath

    def get_export_files(self):
        """Get list of exported files"""
        if os.path.exists(self.export_dir):
            return sorted(os.listdir(self.export_dir), reverse=True)
        return []
